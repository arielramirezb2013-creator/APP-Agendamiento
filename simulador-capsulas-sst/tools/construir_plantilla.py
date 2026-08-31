"""
Reconstruye la hoja «FORMATO FURAT» del formato oficial (furat.xls, F 2015 - PR versión 3)
como vendor/plantilla_furat.xlsx, conservando textos, combinaciones, bordes, rellenos,
fuentes, alineaciones, anchos de columna y altos de fila. La app incrusta ese archivo y
lo diligencia en el navegador sin tocar el formato.

Uso: python3 tools/construir_plantilla.py ruta/al/furat.xls
Requiere: pip install xlrd==1.2.0 openpyxl
"""
import sys, os
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ORIGEN = sys.argv[1] if len(sys.argv) > 1 else "furat.xls"
DESTINO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "vendor", "plantilla_furat.xlsx")

LINEA = {1: "thin", 2: "medium", 3: "dashed", 4: "dotted", 5: "thick", 6: "double", 7: "hair",
         8: "mediumDashed", 9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot", 12: "mediumDashDotDot", 13: "slantDashDot"}
HORIZ = {1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify", 6: "centerContinuous", 7: "distributed"}
VERT = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}

wb_in = xlrd.open_workbook(ORIGEN, formatting_info=True)
ws_in = wb_in.sheet_by_name("FORMATO FURAT")

def color(idx):
    rgb = wb_in.colour_map.get(idx)
    if not rgb or rgb == (0, 0, 0) and idx in (64, 32767):
        return None
    return "FF%02X%02X%02X" % rgb

def lado(estilo, cidx):
    if not estilo:
        return Side()
    return Side(style=LINEA.get(estilo, "thin"), color=color(cidx) or "FF000000")

wb_out = Workbook()
ws = wb_out.active
ws.title = "FORMATO FURAT"
ws.sheet_view.showGridLines = False

for c in range(ws_in.ncols):
    ci = ws_in.colinfo_map.get(c)
    w = (ci.width / 256.0) if ci else ws_in.defcolwidth or 8.43
    ws.column_dimensions[get_column_letter(c + 1)].width = w
for r in range(ws_in.nrows):
    ri = ws_in.rowinfo_map.get(r)
    if ri and ri.height:
        ws.row_dimensions[r + 1].height = ri.height / 20.0

for r in range(ws_in.nrows):
    for c in range(ws_in.ncols):
        cell = ws.cell(row=r + 1, column=c + 1)
        v = ws_in.cell_value(r, c)
        if v not in ("", None):
            if ws_in.cell_type(r, c) == xlrd.XL_CELL_NUMBER and v == int(v):
                v = int(v)
            cell.value = v
        xf = wb_in.xf_list[ws_in.cell_xf_index(r, c)]
        f = wb_in.font_list[xf.font_index]
        cell.font = Font(name=f.name, size=f.height / 20.0, bold=bool(f.weight >= 700), italic=bool(f.italic),
                         underline="single" if f.underline_type else None, color=color(f.colour_index))
        b = xf.border
        cell.border = Border(left=lado(b.left_line_style, b.left_colour_index), right=lado(b.right_line_style, b.right_colour_index),
                             top=lado(b.top_line_style, b.top_colour_index), bottom=lado(b.bottom_line_style, b.bottom_colour_index))
        bg = xf.background
        if bg.fill_pattern == 1:
            fg = color(bg.pattern_colour_index)
            if fg:
                cell.fill = PatternFill(fill_type="solid", fgColor=fg)
        al = xf.alignment
        cell.alignment = Alignment(horizontal=HORIZ.get(al.hor_align), vertical=VERT.get(al.vert_align, "center"),
                                   wrap_text=bool(al.text_wrapped), shrink_to_fit=bool(getattr(al, "shrink_to_fit", 0)))

for (r1, r2, c1, c2) in ws_in.merged_cells:
    ws.merge_cells(start_row=r1 + 1, start_column=c1 + 1, end_row=r2, end_column=c2)

ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
for m, v in (("left", 0.25), ("right", 0.25), ("top", 0.3), ("bottom", 0.3)):
    setattr(ws.page_margins, m, v)

os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
wb_out.save(DESTINO)
print(f"{os.path.relpath(DESTINO)} · {os.path.getsize(DESTINO)/1024:.1f} KB · {ws_in.nrows}×{ws_in.ncols} · {len(ws_in.merged_cells)} combinaciones")
