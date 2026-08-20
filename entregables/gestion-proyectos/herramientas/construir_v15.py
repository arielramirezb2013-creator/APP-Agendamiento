# -*- coding: utf-8 -*-
"""Construye Mapa_de_usuario_y_validacion_Rehavid_v11.xlsx · segunda ronda de validación."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from contenido_v15 import MAPA, ROLES, ORDEN_ROLES, PASOS

DESTINO = "/home/user/APP-Agendamiento/entregables/gestion-proyectos/Historia_de_usuario_Rehavid_v15.xlsx"

FUENTE = "Segoe UI"
MORADO = "4125CF"
MORADO_OSCURO = "2E1A99"
LILA = "EDE9FC"
LILA_CLARO = "F4F2FF"
CREMA = "FFF9E0"
BORDE_COLOR = "C4BEF5"
VERDE = "00C878"

f_titulo = Font(name=FUENTE, size=14, bold=True, color="FFFFFF")
f_sub = Font(name=FUENTE, size=10, color="FFFFFF")
f_head = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
f_body = Font(name=FUENTE, size=10)
f_body_b = Font(name=FUENTE, size=10, bold=True)
f_etiqueta = Font(name=FUENTE, size=10, bold=True, color=MORADO)

fill_titulo = PatternFill("solid", fgColor=MORADO)
fill_verde = PatternFill("solid", fgColor=VERDE)
fill_head = PatternFill("solid", fgColor=MORADO)
fill_sub = PatternFill("solid", fgColor=MORADO_OSCURO)
fill_alt = PatternFill("solid", fgColor=LILA_CLARO)
fill_input = PatternFill("solid", fgColor=CREMA)

lado = Side(style="thin", color=BORDE_COLOR)
borde = Border(left=lado, right=lado, top=lado, bottom=lado)
al_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
al_head = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_centro = Alignment(horizontal="center", vertical="top", wrap_text=True)
al_titulo = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

TITULO_GRAL = ("REHAVID S.A.S. · Gestión de proyectos v15 · Historia de usuario y validación · TERCERA RONDA · salida a producción")
PIE_FUENTE = ("Fuente: Especificación funcional y flujograma del proceso de gestión de proyectos, versión 1.1, Rehavid S.A.S. "
              "Tercera ronda sobre la versión 15 (simplificada a cuatro pestañas, con botón único «Actualizar»). Cada cargo cierra sus hallazgos de la segunda ronda (CORREGIDO o PENDIENTE), revisa su base de alimentación y da su veredicto de salida a producción. "
              "Los pasos 1 a 11 corresponden al numeral 7 de la especificación.")

def poner_titulo(ws, ncols, texto, subtexto=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = f_titulo; c.fill = fill_titulo; c.alignment = al_titulo
    ws.row_dimensions[1].height = 30
    for col in range(1, ncols + 1): ws.cell(row=1, column=col).fill = fill_titulo
    # franja verde de acento, como la app
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    for col in range(1, ncols + 1): ws.cell(row=2, column=col).fill = fill_verde
    ws.row_dimensions[2].height = 4
    if subtexto is None: return 2
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1, value=subtexto)
    c.font = f_sub; c.fill = fill_sub; c.alignment = al_titulo
    for col in range(1, ncols + 1): ws.cell(row=3, column=col).fill = fill_sub
    ws.row_dimensions[3].height = 30
    return 3

def preparar_impresion(ws, fila_encabezado):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "%d:%d" % (fila_encabezado, fila_encabezado)
    ws.page_margins.left = 0.4; ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

def escribir_encabezados(ws, fila, encabezados, anchos):
    for i, (texto, ancho) in enumerate(zip(encabezados, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=texto)
        c.font = f_head; c.fill = fill_head; c.alignment = al_head; c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[fila].height = 34

def pintar_fila(ws, fila, ncols, alterna, cols_input=(), cols_centro=()):
    for col in range(1, ncols + 1):
        c = ws.cell(row=fila, column=col)
        c.font = f_body; c.border = borde
        c.alignment = al_centro if col in cols_centro else al_wrap
        if col in cols_input: c.fill = fill_input
        elif alterna: c.fill = fill_alt

# ============================ PESTAÑA 1 · MAPA ============================
ws = wb.active
ws.title = "Historia de usuario"
ENC = ["Rol / cargo", "Épica", "Qué hace",
       "Módulo donde lo hace", "Dato o decisión que produce", "Qué necesita para poder hacerlo"]
ANCHOS = [24, 32, 62, 30, 46, 50]
NC = len(ENC)
poner_titulo(ws, NC, TITULO_GRAL,
    "Pestaña 1 · Historia de usuario de la versión 15. Cada fila es una intervención real de un cargo en las épicas del "
    "flujo: alimentación por bases, venta, activación, asignación, negociación, requisitos, plan final, ejecución con el "
    "botón «Actualizar», cierres y veredicto. Solo lectura: esta pestaña no se diligencia.")
FILA_ENC = 5
escribir_encabezados(ws, FILA_ENC, ENC, ANCHOS)
fila = FILA_ENC + 1
for paso, rol, hace, modulo, produce, necesita in sorted(MAPA, key=lambda r: (r[0], r[1])):
    ws.cell(row=fila, column=1, value=rol)
    ws.cell(row=fila, column=2, value=PASOS[paso])
    ws.cell(row=fila, column=3, value=hace)
    ws.cell(row=fila, column=4, value=modulo)
    ws.cell(row=fila, column=5, value=produce)
    ws.cell(row=fila, column=6, value=necesita)
    pintar_fila(ws, fila, NC, alterna=(paso % 2 == 0))
    ws.cell(row=fila, column=1).font = f_body_b
    fila += 1
ws.freeze_panes = "A%d" % (FILA_ENC + 1)
ws.auto_filter.ref = "A%d:%s%d" % (FILA_ENC, get_column_letter(NC), fila - 1)
preparar_impresion(ws, FILA_ENC)
fila += 1
ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=NC)
c = ws.cell(row=fila, column=1, value=PIE_FUENTE)
c.font = Font(name=FUENTE, size=9, italic=True, color="55557A"); c.alignment = al_wrap
ws.sheet_view.showGridLines = False
FILAS_MAPA = len(MAPA)

# ==================== PESTAÑA 2 · DIAGRAMA DEL PROCESO ====================
from openpyxl.drawing.image import Image as XLImage
ws = wb.create_sheet(title="Flujo del app", index=1)
NCD = 12
poner_titulo(ws, NCD, TITULO_GRAL,
    "Pestaña 2 · El flujo que hace el app: la alimentación por bases y los once pasos; por cada paso, qué se ingresa, "
    "qué opera el aplicativo y qué sale. Solo lectura.")
for col in range(1, NCD + 1):
    ws.column_dimensions[get_column_letter(col)].width = 13
ws.row_dimensions[4].height = 8
img = XLImage("diagrama_v15.png")
# El PNG está a 200 ppp; se presenta a la mitad de sus píxeles para verse nítido y proporcionado.
import PIL.Image as _PIL
_im = _PIL.open("diagrama_v15.png")
img.width = _im.size[0] // 2
img.height = _im.size[1] // 2
ws.add_image(img, "B5")
ws.sheet_view.showGridLines = False
preparar_impresion(ws, 1)
ws.page_setup.orientation = "portrait"

# ============================ PESTAÑAS POR CARGO ============================
ENC_ROL = ["N°", "Épica", "Actividad / tarea",
           "Paso a paso que debe seguir en la aplicación", "Resultado esperado",
           "¿Funciona? (Sí / No / Parcial)", "Claridad (1-5)", "Observaciones", "Qué cambiaría"]
ANCHOS_ROL = [6, 28, 40, 76, 54, 22, 14, 58, 58]
NCR = len(ENC_ROL)
COLS_INPUT = (6, 7, 8, 9)

INSTRUCCIONES = ("Instrucciones · SEGUNDA RONDA sobre la versión 12. Su usuario en la aplicación es su CARGO (selector superior "
                 "derecho); ya no hay nombres de persona. Recorra cada actividad en orden, marque si funciona, califique de 1 a 5 la "
                 "claridad y escriba sus observaciones. Diligencie únicamente las cuatro columnas en color crema.")
EJEMPLO = ("Cómo diligenciar · Ejemplo: ¿Funciona? = Parcial · Claridad = 3 · Observaciones = «La contrapropuesta no muestra la "
           "fecha anterior» · Qué cambiaría = «Mostrar fecha anterior y nueva lado a lado». Las listas de ¿Funciona? y Claridad "
           "se abren con la flecha de la celda.")

conteo = {}
for nombre_rol, nombre_pestana in ORDEN_ROLES:
    datos = ROLES[nombre_rol]
    ws = wb.create_sheet(title=nombre_pestana[:31])
    poner_titulo(ws, NCR, TITULO_GRAL + " · Cargo: " + datos["nombre"])
    def bloque(fila, etiqueta, texto, alto=None):
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=NCR)
        c = ws.cell(row=fila, column=1, value=("%s: %s" % (etiqueta, texto)) if etiqueta else texto)
        c.font = f_body; c.alignment = al_wrap
        if alto: ws.row_dimensions[fila].height = alto
        return c
    c = bloque(3, "Cargo", datos["nombre"], 22)
    c.font = Font(name=FUENTE, size=12, bold=True, color=MORADO)
    bloque(4, "Responsabilidad en la versión 12", datos["responsabilidad"], 28)
    bloque(5, None, INSTRUCCIONES, 30)
    bloque(6, None, EJEMPLO, 28)
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    c = ws.cell(row=7, column=1, value="Nombre de quien revisa:")
    c.font = f_etiqueta; c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=5)
    for col in (3, 4, 5):
        cc = ws.cell(row=7, column=col); cc.fill = fill_input; cc.border = borde
    ws.cell(row=7, column=3).font = f_body
    c = ws.cell(row=7, column=6, value="Fecha de la revisión:")
    c.font = f_etiqueta; c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=7, start_column=7, end_row=7, end_column=8)
    for col in (7, 8):
        cc = ws.cell(row=7, column=col); cc.fill = fill_input; cc.border = borde
    cf = ws.cell(row=7, column=7); cf.font = f_body; cf.number_format = "DD/MM/YYYY"
    cf.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 8

    FILA_ENC = 9
    escribir_encabezados(ws, FILA_ENC, ENC_ROL, ANCHOS_ROL)
    fila = FILA_ENC + 1
    for i, (paso, actividad, pasoapaso, esperado) in enumerate(sorted(datos["filas"], key=lambda f: f[0]), start=1):
        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=PASOS[paso])
        ws.cell(row=fila, column=3, value=actividad)
        ws.cell(row=fila, column=4, value=pasoapaso)
        ws.cell(row=fila, column=5, value=esperado)
        pintar_fila(ws, fila, NCR, alterna=(i % 2 == 0), cols_input=COLS_INPUT, cols_centro=(1, 6, 7))
        fila += 1
    ultima = fila - 1
    conteo[nombre_pestana] = len(datos["filas"])

    dv_f = DataValidation(type="list", formula1='"Sí,No,Parcial,No aplica"', allow_blank=True, showDropDown=False)
    dv_f.errorTitle = "Valor no válido"; dv_f.error = "Seleccione: Sí, No, Parcial o No aplica."
    dv_f.promptTitle = "¿Funciona?"; dv_f.prompt = "Seleccione: Sí, No, Parcial o No aplica."
    ws.add_data_validation(dv_f); dv_f.add("F%d:F%d" % (FILA_ENC + 1, ultima))
    dv_c = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True, showDropDown=False)
    dv_c.errorTitle = "Valor no válido"; dv_c.error = "Califique con un entero de 1 a 5."
    dv_c.promptTitle = "Claridad (1-5)"; dv_c.prompt = "1 = nada clara · 5 = totalmente clara."
    ws.add_data_validation(dv_c); dv_c.add("G%d:G%d" % (FILA_ENC + 1, ultima))

    ws.freeze_panes = "C%d" % (FILA_ENC + 1)
    ws.auto_filter.ref = "A%d:%s%d" % (FILA_ENC, get_column_letter(NCR), ultima)
    preparar_impresion(ws, FILA_ENC)
    pie = ultima + 2
    ws.merge_cells(start_row=pie, start_column=1, end_row=pie, end_column=NCR)
    c = ws.cell(row=pie, column=1, value=PIE_FUENTE + " Traslade cada observación a la pestaña «Resumen de observaciones» con el N° de la fila.")
    c.font = Font(name=FUENTE, size=9, italic=True, color="55557A"); c.alignment = al_wrap
    ws.sheet_view.showGridLines = False

# ==================== PESTAÑA FINAL · RESUMEN ====================
ws = wb.create_sheet(title="Resumen de observaciones")
ENC_RES = ["Cargo", "N° de la observación", "Módulo", "Severidad", "Observación",
           "Decisión", "Responsable", "Fecha compromiso"]
ANCHOS_RES = [28, 18, 34, 18, 66, 60, 26, 18]
NCS = len(ENC_RES)
poner_titulo(ws, NCS, TITULO_GRAL,
    "Pestaña final · Consolidación de las observaciones de la tercera ronda y DECISIÓN DE SALIDA A PRODUCCIÓN. El N° corresponde a la pestaña del cargo.")
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=NCS)
c = ws.cell(row=4, column=1, value=("Instrucciones · Traslade aquí cada observación. Severidad: Bloquea (impide continuar el flujo), "
                                    "Alta, Media, Baja o Mejora. Cargo, Módulo y Severidad tienen lista desplegable; la fecha va en DD/MM/AAAA."))
c.font = f_body; c.alignment = al_wrap
ws.row_dimensions[4].height = 34
FILA_ENC = 6
escribir_encabezados(ws, FILA_ENC, ENC_RES, ANCHOS_RES)
fill_input_alt = PatternFill("solid", fgColor="FFFDF3")
N_VACIAS = 60
for i in range(N_VACIAS):
    fila = FILA_ENC + 1 + i
    relleno = fill_input if i % 2 == 0 else fill_input_alt
    for col in range(1, NCS + 1):
        c = ws.cell(row=fila, column=col)
        c.font = f_body; c.border = borde
        c.alignment = al_centro if col in (2, 4, 8) else al_wrap
        c.fill = relleno
    ws.cell(row=fila, column=8).number_format = "DD/MM/YYYY"
ultima = FILA_ENC + N_VACIAS
cargos = ",".join([r[0] for r in ORDEN_ROLES])
dv_rol = DataValidation(type="list", formula1='"%s"' % cargos, allow_blank=True, showDropDown=False)
dv_rol.promptTitle = "Cargo"; dv_rol.prompt = "Seleccione el cargo que registró la observación."
ws.add_data_validation(dv_rol); dv_rol.add("A%d:A%d" % (FILA_ENC + 1, ultima))
dv_sev = DataValidation(type="list", formula1='"Bloquea,Alta,Media,Baja,Mejora"', allow_blank=True, showDropDown=False)
dv_sev.errorTitle = "Valor no válido"; dv_sev.error = "Seleccione: Bloquea, Alta, Media, Baja o Mejora."
dv_sev.promptTitle = "Severidad"; dv_sev.prompt = "Bloquea · Alta · Media · Baja · Mejora"
ws.add_data_validation(dv_sev); dv_sev.add("D%d:D%d" % (FILA_ENC + 1, ultima))
modulos = ("Guía,Resumen,Actividades (tablero),Actividades (lista),Mi trabajo,Botón Actualizar,Calendario,"
           "Ficha del proyecto,Crear (venta),Parametrización,Bases y auditoría,Base de Excel,Varios módulos")
dv_mod = DataValidation(type="list", formula1='"%s"' % modulos, allow_blank=True, showDropDown=False)
dv_mod.promptTitle = "Módulo"; dv_mod.prompt = "Módulo donde se detectó la observación."
ws.add_data_validation(dv_mod); dv_mod.add("C%d:C%d" % (FILA_ENC + 1, ultima))
ws.freeze_panes = "A%d" % (FILA_ENC + 1)
ws.auto_filter.ref = "A%d:%s%d" % (FILA_ENC, get_column_letter(NCS), ultima)

# ---- Decisión de salida a producción ----
fila_dec = ultima + 3
ws.merge_cells(start_row=fila_dec, start_column=1, end_row=fila_dec, end_column=NCS)
cd = ws.cell(row=fila_dec, column=1, value="DECISIÓN DE SALIDA A PRODUCCIÓN · una fila por cargo")
cd.font = f_titulo; cd.fill = fill_titulo; cd.alignment = al_titulo
for col in range(1, NCS + 1): ws.cell(row=fila_dec, column=col).fill = fill_titulo
ws.row_dimensions[fila_dec].height = 26
fila_dec += 1
ws.merge_cells(start_row=fila_dec, start_column=1, end_row=fila_dec, end_column=NCS)
cd = ws.cell(row=fila_dec, column=1, value=("Al terminar su pestaña, cada cargo marca su veredicto: la versión sale a producción "
    "solo si ningún cargo responde «No» y las condiciones quedan acordadas con la dirección."))
cd.font = f_body; cd.alignment = al_wrap
ws.row_dimensions[fila_dec].height = 30
fila_dec += 1
ENC_DEC = ["Cargo", "¿Aprueba la salida a producción?", "Condiciones o ajustes exigidos", "Nombre de quien valida", "Fecha"]
for i, (texto, ancho) in enumerate(zip(ENC_DEC, [28, 30, 80, 30, 18]), start=1):
    cdd = ws.cell(row=fila_dec, column=i, value=texto)
    cdd.font = f_head; cdd.fill = fill_head; cdd.alignment = al_head; cdd.border = borde
ws.row_dimensions[fila_dec].height = 30
dv_dec = DataValidation(type="list", formula1='"Sí, aprobado,Con ajustes menores,No — requiere otra ronda"', allow_blank=True, showDropDown=False)
dv_dec.promptTitle = "Veredicto"; dv_dec.prompt = "Sí, aprobado · Con ajustes menores · No — requiere otra ronda"
ws.add_data_validation(dv_dec)
for j, (rol, _) in enumerate(ORDEN_ROLES, start=1):
    f = fila_dec + j
    ws.cell(row=f, column=1, value=rol).font = f_body_b
    for col in range(1, 6):
        cdd = ws.cell(row=f, column=col)
        cdd.border = borde; cdd.alignment = al_wrap
        if col >= 2: cdd.fill = fill_input
    ws.cell(row=f, column=5).number_format = "DD/MM/YYYY"
    ws.row_dimensions[f].height = 26
dv_dec.add("B%d:B%d" % (fila_dec + 1, fila_dec + len(ORDEN_ROLES)))
preparar_impresion(ws, FILA_ENC)
pie = ultima + 2
ws.merge_cells(start_row=pie, start_column=1, end_row=pie, end_column=NCS)
c = ws.cell(row=pie, column=1, value=PIE_FUENTE)
c.font = Font(name=FUENTE, size=9, italic=True, color="55557A"); c.alignment = al_wrap
ws.sheet_view.showGridLines = False

wb.properties.title = "Historia de usuario y validación v15 · Rehavid · Gestión de proyectos · tercera ronda"
wb.properties.subject = "Especificación funcional versión 1.1 · usuarios por cargo"
wb.properties.creator = "Rehavid S.A.S."
wb.properties.language = "es-CO"
os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
wb.save(DESTINO)
print("Guardado:", DESTINO)
print("Mapa de usuario:", FILAS_MAPA, "filas")
for k, v in conteo.items(): print("  %-30s %d actividades" % (k, v))
