# -*- coding: utf-8 -*-
"""Genera las tres bases de alimentación de Rehavid · Gestión de proyectos v16 (versión final).
1. Base_BI.xlsx — productos de Tablero BI (Gestor de Datos)
2. Base_Diseno_Ergonomico.xlsx — productos de Diseño (Diseñador)
3. Base_General.xlsx — mediciones, cápsulas y demás, más los catálogos.
Las tres comparten el mismo formato de hojas; la app las carga desde
la barra lateral → Bases y auditoría → Cargar base de alimentación (Excel)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEST = "/home/user/APP-Agendamiento/entregables/gestion-proyectos/bases"
os.makedirs(DEST, exist_ok=True)

FUENTE = "Segoe UI"
MORADO, MORADO_O, LILA, LILA_C = "4125CF", "2E1A99", "EDE9FC", "F4F2FF"
VERDE, CREMA, BORDE_C = "00C878", "FFF9E0", "C4BEF5"

f_tit  = Font(name=FUENTE, size=13, bold=True, color="FFFFFF")
f_sub  = Font(name=FUENTE, size=9.5, color="FFFFFF")
f_head = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
f_body = Font(name=FUENTE, size=10)
f_hint = Font(name=FUENTE, size=8.5, italic=True, color="55557A")
fill_tit  = PatternFill("solid", fgColor=MORADO)
fill_sub  = PatternFill("solid", fgColor=MORADO_O)
fill_head = PatternFill("solid", fgColor=MORADO)
fill_verde= PatternFill("solid", fgColor=VERDE)
fill_alt  = PatternFill("solid", fgColor=LILA_C)
fill_in   = PatternFill("solid", fgColor=CREMA)
lado  = Side(style="thin", color=BORDE_C)
borde = Border(left=lado, right=lado, top=lado, bottom=lado)
al_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
al_head = Alignment(horizontal="left", vertical="center", wrap_text=True)

CARGOS_TXT = "CEO, Gerente, Director de Proyectos, Jefe de Operaciones, Supervisor Contactabilidad, Gestor de Datos 1, Gestor de Datos 2, Diseñador"

# ---------- definición de hojas comunes ----------
# (nombre, [(encabezado, ancho, pista)], [filas ejemplo], validaciones {col: formula})
def hojas_comunes(ejemplos):
    return [
      ("Proyectos", [
        ("codigo", 12, "Opcional. Vacío = la app lo genera (PRJ-…)."),
        ("empresa", 26, "Obligatoria. Si no existe, la app la crea en el catálogo."),
        ("nit", 14, "Opcional."),
        ("ppr", 12, "Obligatorio. Código del catálogo (PPR-DME, PPR-ERG…)."),
        ("tipo", 10, "Tipo 1 o Tipo 2."),
        ("orden_servicio", 16, "Opcional."),
        ("fecha_venta", 14, "AAAA-MM-DD o fecha de Excel."),
        ("valor_facturado", 16, "Número, sin puntos de miles."),
        ("plan_estimado", 44, "Obligatorio para activar."),
        ("observaciones_venta", 44, "Obligatorias para activar (mínimo 10 caracteres)."),
        ("activar", 10, "SI = la app corre las validaciones y activa; si alguna falla queda en Borrador."),
      ], ejemplos["Proyectos"], {"E": '"Tipo 1,Tipo 2"', "K": '"SI,NO"'}),
      ("Productos", [
        ("proyecto", 12, "Código de la hoja Proyectos."),
        ("producto", 34, "Nombre del producto vendido."),
        ("categoria", 18, "Medición objetiva · Tablero BI · Diseño · Cápsulas · Otro. Define el cargo responsable."),
        ("horas_vendidas", 14, "Obligatorias (> 0)."),
        ("valor_hora", 12, "Número."),
        ("tecnologia", 16, "Solo mediciones: Xsens, TuMeke, BVB, Tobii 3…"),
        ("complejidad", 12, "Solo diseño: Baja, Media o Alta."),
        ("personas_a_impactar", 18, "Solo cápsulas."),
        ("fecha_solicitada", 15, "Obligatoria para activar: compromiso preliminar de la venta."),
        ("inicio", 12, "Opcional."),
        ("fecha_objetivo", 14, "Opcional."),
        ("avance_pct", 11, "0 a 100. Solo si el producto ya viene andando."),
        ("horas_ejecutadas", 15, "Solo si ya viene andando."),
        ("notas", 30, "Opcional."),
        ("responsable", 24, "Opcional. Uno de los ocho cargos; si va vacío, asigna el motor."),
      ], ejemplos["Productos"], {"C": '"Medición objetiva,Tablero BI,Diseño,Cápsulas,Otro"', "G": '"Baja,Media,Alta"'}),
      ("Contactos", [
        ("proyecto", 12, "Código de la hoja Proyectos."),
        ("nombre", 24, "Obligatorio."),
        ("cargo", 22, "Cargo en la empresa cliente."),
        ("correo", 30, "Obligatorio para el contacto principal."),
        ("telefono", 14, "Obligatorio para el contacto principal."),
        ("principal", 10, "SI en exactamente un contacto por proyecto."),
      ], ejemplos["Contactos"], {"F": '"SI,NO"'}),
      ("Hitos", [
        ("proyecto", 12, "Código de la hoja Proyectos."),
        ("producto", 34, "Nombre exacto de la hoja Productos."),
        ("hito", 34, "Nombre del hito o sesión."),
        ("fecha", 12, "AAAA-MM-DD o fecha de Excel."),
        ("avance_pct", 11, "0 a 100."),
        ("horas_asignadas", 15, "Número."),
        ("horas_ejecutadas", 15, "Número."),
      ], ejemplos["Hitos"], {}),
      ("Seguimiento", [
        ("proyecto", 12, "Código de la hoja Proyectos."),
        ("producto", 34, "Nombre exacto de la hoja Productos."),
        ("fecha", 12, "AAAA-MM-DD o fecha de Excel."),
        ("tipo", 14, "Avance · Retraso · Suspensión · Bloqueo · Reanudación · Reprogramación · Nota."),
        ("descripcion", 44, "Qué pasó (causa)."),
        ("accion", 34, "Acción correctiva, si aplica."),
        ("responsable", 22, "Uno de los ocho cargos."),
        ("nueva_fecha", 12, "Solo retraso, bloqueo o reprogramación."),
        ("avance_pct", 11, "0 a 100, si cambió."),
        ("horas_ejecutadas", 15, "Si cambiaron."),
      ], ejemplos["Seguimiento"], {"D": '"Avance,Retraso,Suspensión,Bloqueo,Reanudación,Reprogramación,Nota"'}),
    ]

HOJAS_CATALOGO = [
  ("Empresas", [("empresa", 28, "Razón social."), ("nit", 16, "Opcional."), ("notas", 34, "Opcional.")], [
    ["Alimentos del Norte S.A.", "900123456-1", "Cliente activo"],
    ["Agroindustrias La Sabana S.A.S.", "901777333-5", "Cliente nuevo 2026"],
    ["Curtiembres El Cedro S.A.S.", "901555222-3", "Cliente nuevo 2026"]], {}),
  ("PPR", [("codigo", 12, "PPR-…"), ("descripcion", 34, "Programa interno.")], [
    ["PPR-DME", "Desórdenes musculoesqueléticos"], ["PPR-ERG", "Ergonomía"]], {}),
  ("Productos y tarifas", [("producto", 34, "Nombre de catálogo."), ("categoria", 18, "Categoría."),
                           ("horas", 10, "Costo en horas."), ("valor_hora", 12, "Tarifa.")], [
    ["Mediciones objetivas de carga física", "Medición objetiva", 120, 130000],
    ["Auditoría de puestos con video", "Otro", 50, 98000]],
   {"B": '"Medición objetiva,Tablero BI,Diseño,Cápsulas,Otro"'}),
  ("Ordenes de servicio", [("orden", 16, "Referencia comercial."), ("empresa", 28, "Cliente."),
                           ("fecha", 12, "AAAA-MM-DD.")], [
    ["OS-2026-140", "Curtiembres El Cedro S.A.S.", "2026-08-20"]], {}),
]

def escribir_hoja(wb, nombre, cols, filas, valids, es_primera=False):
    ws = wb.create_sheet(title=nombre) if wb.worksheets and not es_primera else wb.active
    if es_primera: ws.title = nombre
    NC = len(cols)
    # título + franja + pista
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c = ws.cell(row=1, column=1, value=f"BASE DE ALIMENTACIÓN · {nombre.upper()}")
    c.font = f_tit; c.fill = fill_tit; c.alignment = al_head
    for col in range(1, NC + 1): ws.cell(row=1, column=col).fill = fill_tit
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    for col in range(1, NC + 1): ws.cell(row=2, column=col).fill = fill_verde
    ws.row_dimensions[2].height = 3
    # encabezados (fila 3) + pistas (fila 4)
    for i, (h, w, hint) in enumerate(cols, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = f_head; c.fill = fill_head; c.alignment = al_head; c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = w
        c2 = ws.cell(row=4, column=i, value=hint)
        c2.font = f_hint; c2.alignment = al_wrap; c2.border = borde
        c2.fill = PatternFill("solid", fgColor=LILA)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 34
    # ejemplos
    for r, fila in enumerate(filas, start=5):
        for i, v in enumerate(fila, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = f_body; c.border = borde; c.alignment = al_wrap
            if (r - 5) % 2 == 1: c.fill = fill_alt
    # filas vacías listas para diligenciar
    for r in range(5 + len(filas), 5 + len(filas) + 30):
        for i in range(1, NC + 1):
            c = ws.cell(row=r, column=i)
            c.border = borde; c.font = f_body; c.fill = fill_in if (r % 2 == 0) else PatternFill("solid", fgColor="FFFDF3")
    ultima = 5 + len(filas) + 29
    for colL, formula in valids.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{colL}5:{colL}{ultima}")
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws

def hoja_instrucciones(wb, titulo, alcance, hojas_nombres):
    ws = wb.active
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    ws.merge_cells("B2:B2")
    c = ws.cell(row=2, column=2, value=titulo)
    c.font = Font(name=FUENTE, size=15, bold=True, color=MORADO)
    filas = [
      "",
      alcance,
      "",
      "CÓMO SE USA",
      "1. Diligencie las hojas en este orden: " + " → ".join(hojas_nombres) + ".",
      "2. Las dos primeras filas de cada hoja son el encabezado y una pista de cada columna; no las borre ni las mueva.",
      "3. Las filas de ejemplo son una guía: reemplácelas o bórrelas antes de la carga real.",
      "4. Las fechas van como AAAA-MM-DD o como fecha normal de Excel; ambas se leen bien.",
      "5. En «activar» escriba SI solo cuando la fila tenga todo lo obligatorio; la app corre las mismas",
      "   validaciones del registro de venta y, si alguna falla, el proyecto queda en Borrador con el detalle.",
      "6. Guarde el archivo como .xlsx (formato normal de Excel).",
      "",
      "CÓMO SE CARGA EN LA APLICACIÓN",
      "La carga la hace un ADMINISTRADOR (CEO, Gerente, Director de Proyectos o Jefe de Operaciones):",
      "barra lateral → Bases y auditoría → «Cargar base de alimentación (Excel)» → elija este archivo. La carga NO borra lo existente: crea lo nuevo y",
      "actualiza lo que coincida por código de proyecto y nombre de producto, y muestra un resumen de lo cargado.",
      "",
      "REGLAS QUE APLICA LA CARGA",
      "· El cargo responsable lo asigna el motor según la categoría, salvo que la columna «responsable» traiga",
      "  uno de los ocho cargos (sirve, por ejemplo, para elegir entre Gestor de Datos 1 y 2):",
      "  Medición objetiva → Director de Proyectos · Tablero BI → Gestor de Datos 1 · Diseño → Diseñador ·",
      "  Cápsulas → Supervisor Contactabilidad · Otro → coordinación del tipo.",
      "· Los usuarios de la aplicación son los ocho cargos: " + CARGOS_TXT + ".",
      "· Empresas, PPR y órdenes que no existan se crean en el catálogo automáticamente.",
      "· El seguimiento se registra como novedades en la bitácora del producto, con su tipo y su fecha.",
    ]
    r = 4
    for t in filas:
        c = ws.cell(row=r, column=2, value=t)
        if t in ("CÓMO SE USA", "CÓMO SE CARGA EN LA APLICACIÓN", "REGLAS QUE APLICA LA CARGA"):
            c.font = Font(name=FUENTE, size=11, bold=True, color=MORADO)
        else:
            c.font = f_body
        c.alignment = al_wrap
        r += 1
    ws.sheet_view.showGridLines = False

def construir_base(nombre_archivo, titulo, alcance, ejemplos, con_catalogos):
    wb = Workbook()
    hoja_instrucciones(wb, titulo, alcance, [h[0] for h in hojas_comunes(ejemplos)])
    for nombre, cols, filas, valids in hojas_comunes(ejemplos):
        escribir_hoja(wb, nombre, cols, filas, valids)
    if con_catalogos:
        for nombre, cols, filas, valids in HOJAS_CATALOGO:
            escribir_hoja(wb, nombre, cols, filas, valids)
    wb.properties.title = titulo
    wb.properties.creator = "Rehavid S.A.S."
    wb.properties.language = "es-CO"
    ruta = os.path.join(DEST, nombre_archivo)
    wb.save(ruta)
    print("OK", ruta)

# ---------- ejemplos por base ----------
EJ_BI = {
  "Proyectos": [
    ["BI-001", "Curtiembres El Cedro S.A.S.", "901555222-3", "PPR-ANA", "Tipo 2", "OS-2026-140", "2026-08-05",
     0, "Tablero BI de ausentismo y rotación con corte semanal, publicado en el micrositio.",
     "El cliente entrega los históricos en la primera semana. Corte semanal los lunes. Acceso de solo lectura al ERP.", "SI"],
  ],
  "Productos": [
    ["BI-001", "Tablero BI de ausentismo", "Tablero BI", 60, 110000, "", "", "", "2026-10-15", "2026-08-25", "2026-10-15", 10, 6, "Con actualización semanal"],
    ["BI-001", "Tablero BI de indicadores DME", "Tablero BI", 45, 110000, "", "", "", "2026-11-10", "", "", "", "", "", "Gestor de Datos 2"],
  ],
  "Contactos": [
    ["BI-001", "Laura Méndez", "Directora de Gestión Humana", "laura.mendez@elcedro.demo", "3125550011", "SI"],
    ["BI-001", "Óscar Pinto", "Analista de nómina", "oscar.pinto@elcedro.demo", "3125550022", "NO"],
  ],
  "Hitos": [
    ["BI-001", "Tablero BI de ausentismo", "Recepción de históricos", "2026-09-01", 100, 8, 8],
    ["BI-001", "Tablero BI de ausentismo", "Modelo de datos y medidas", "2026-09-22", 20, 24, 5],
    ["BI-001", "Tablero BI de ausentismo", "Publicación y capacitación", "2026-10-15", 0, 28, 0],
  ],
  "Seguimiento": [
    ["BI-001", "Tablero BI de ausentismo", "2026-09-02", "Avance", "Históricos recibidos completos y validados.", "", "Gestor de Datos 1", "", 10, 6],
  ],
}
EJ_DIS = {
  "Proyectos": [
    ["DIS-001", "Metalúrgica Andina S.A.S.", "", "PPR-DIS", "Tipo 2", "OS-2026-131", "2026-08-01",
     0, "Rediseño del puesto de corte con validación en sitio y ficha técnica.",
     "El inicio depende de la parada de mantenimiento de septiembre. Rehavid entrega planos y ficha técnica.", "NO"],
  ],
  "Productos": [
    ["DIS-001", "Rediseño del puesto de corte", "Diseño", 90, 120000, "", "Alta", "", "2026-11-28", "", "", "", "", "Depende de la parada de planta"],
    ["DIS-001", "Rediseño de la estación de empaque", "Diseño", 45, 120000, "", "Media", "", "2026-12-10", "", "", "", "", ""],
  ],
  "Contactos": [
    ["DIS-001", "Marta Gil", "Gerente de operaciones", "marta.gil@metalandina.demo", "3208889090", "SI"],
  ],
  "Hitos": [
    ["DIS-001", "Rediseño del puesto de corte", "Levantamiento en sitio", "2026-09-20", 0, 30, 0],
    ["DIS-001", "Rediseño del puesto de corte", "Propuesta de diseño", "2026-10-25", 0, 40, 0],
    ["DIS-001", "Rediseño del puesto de corte", "Validación con el cliente", "2026-11-28", 0, 20, 0],
  ],
  "Seguimiento": [
    ["DIS-001", "Rediseño del puesto de corte", "2026-08-28", "Nota", "El cliente confirmó la parada de mantenimiento para la tercera semana de septiembre.", "", "Diseñador", "", "", ""],
  ],
}
EJ_GEN = {
  "Proyectos": [
    ["GEN-001", "Agroindustrias La Sabana S.A.S.", "901777333-5", "PPR-DME", "Tipo 2", "OS-2026-150", "2026-08-10",
     0, "Mediciones objetivas en la línea de empaque y cápsulas para líderes.",
     "Turnos de noche en la sede sur. El informe se presenta al comité del cliente cada quincena.", "SI"],
  ],
  "Productos": [
    ["GEN-001", "Mediciones objetivas línea de empaque", "Medición objetiva", 100, 130000, "Xsens", "", "", "2026-10-30", "2026-09-01", "2026-10-30", "", "", ""],
    ["GEN-001", "Cápsulas para líderes de línea", "Cápsulas", 40, 95000, "", "", 150, "2026-11-15", "", "", "", "", ""],
    ["GEN-001", "Cápsulas de refuerzo para supervisores", "Cápsulas", 20, 95000, "", "", 80, "2026-12-01", "", "", "", "", ""],
    ["GEN-001", "Acompañamiento a la implementación", "Otro", 30, 105000, "", "", "", "2026-12-10", "", "", "", "", ""],
  ],
  "Contactos": [
    ["GEN-001", "Ana Restrepo", "Jefe de SST", "ana.restrepo@lasabana.demo", "3105558877", "SI"],
    ["GEN-001", "Carlos Díaz", "Coordinador de planta", "carlos.diaz@lasabana.demo", "6014447788", "NO"],
  ],
  "Hitos": [
    ["GEN-001", "Mediciones objetivas línea de empaque", "Sesión sede norte", "2026-09-15", 0, 40, 0],
    ["GEN-001", "Mediciones objetivas línea de empaque", "Sesión sede sur", "2026-10-06", 0, 35, 0],
    ["GEN-001", "Mediciones objetivas línea de empaque", "Informe de resultados", "2026-10-30", 0, 25, 0],
  ],
  "Seguimiento": [
    ["GEN-001", "Mediciones objetivas línea de empaque", "2026-09-02", "Nota", "Confirmado el acceso nocturno a la sede sur.", "", "Director de Proyectos", "", "", ""],
  ],
}

construir_base("Base_BI.xlsx",
  "Base de alimentación · BI (Tablero BI)",
  "Alimenta los productos de TABLERO BI, que el motor asigna al cargo Gestor de Datos 1. Use la columna categoria = «Tablero BI».",
  EJ_BI, con_catalogos=False)
construir_base("Base_Diseno_Ergonomico.xlsx",
  "Base de alimentación · Diseño Ergonómico",
  "Alimenta los productos de DISEÑO, que el motor asigna al cargo Diseñador. Use la columna categoria = «Diseño» y diligencie la complejidad.",
  EJ_DIS, con_catalogos=False)
construir_base("Base_General.xlsx",
  "Base de alimentación · General (mediciones, cápsulas y demás) + catálogos",
  "Alimenta los productos de MEDICIÓN OBJETIVA, CÁPSULAS y OTROS, y además los catálogos (empresas, PPR, productos y tarifas, órdenes de servicio).",
  EJ_GEN, con_catalogos=True)
