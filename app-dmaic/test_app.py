#!/usr/bin/env python3
"""
Arnés de pruebas headless de la app Rehavid LSS DMAIC.

Carga el HTML autónomo en Chromium, recorre TODAS las herramientas, escribe
datos, ejercita el motor de exportación y valida los .xlsx/.xlsm resultantes
con openpyxl y olevba.

  python3 test_app.py
"""
import os, sys, json, glob, zipfile, tempfile, traceback
from playwright.sync_api import sync_playwright

CHROME = '/opt/pw-browsers/chromium-1194/chrome-linux/chrome'
HTML = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'REHAVID_LSS_DMAIC.html'))
DL = tempfile.mkdtemp(prefix='rvdl_')

errores, avisos = [], []


def run():
    with sync_playwright() as pw:
        b = pw.chromium.launch(executable_path=CHROME, args=['--no-sandbox'])
        ctx = b.new_context(accept_downloads=True, viewport={'width': 1600, 'height': 1000})
        pg = ctx.new_page()

        pg.on('pageerror', lambda e: errores.append('PAGEERROR: %s' % e))
        pg.on('console', lambda m: (errores.append('CONSOLE.ERROR: %s' % m.text)
                                    if m.type == 'error' else
                                    (avisos.append(m.text) if m.type == 'warning' else None)))

        pg.goto('file://' + HTML)
        pg.wait_for_timeout(1200)

        # ── 1. arranque ────────────────────────────────────────────────
        n = pg.evaluate('typeof TOOLS !== "undefined" ? TOOLS.length : -1')
        print('Herramientas registradas: %s' % n)
        if n < 53:
            errores.append('Se esperaban 53 herramientas, hay %s' % n)

        mods = pg.evaluate('MODULES.map(m=>[m.id, TOOLS.filter(t=>t.mod===m.id).length])')
        for mid, c in mods:
            print('   %-10s %2d formatos' % (mid, c))

        # subsistemas presentes
        for g in ['CHART', 'DIAG', 'ENGINE', 'BI', 'ZIP', 'XL', 'GEN', 'R', 'UI', 'ST']:
            if not pg.evaluate('typeof %s !== "undefined"' % g):
                errores.append('Falta el objeto global %s' % g)

        # ── 2. recorrer TODAS las herramientas ─────────────────────────
        ids = pg.evaluate('TOOLS.map(t=>t.id)')
        rotas = []
        for tid in ids:
            try:
                pg.evaluate('''(id)=>{ const t=TOOL_BY_ID[id]; go('tool', t.mod, t.id); }''', tid)
                pg.wait_for_timeout(45)
                html = pg.eval_on_selector('#view', 'e=>e.innerHTML')
                if 'Error al dibujar' in html:
                    rotas.append(tid)
            except Exception as ex:
                rotas.append('%s (%s)' % (tid, ex))
        if rotas:
            errores.append('Herramientas que fallan al dibujar: %s' % rotas)
        print('Herramientas recorridas: %d · con error: %d' % (len(ids), len(rotas)))

        # ── 3. otras vistas ────────────────────────────────────────────
        for v in ['inicio', 'bi', 'exportar', 'proyectos', 'ajustes']:
            pg.evaluate('(v)=>go(v)', v)
            pg.wait_for_timeout(160)
            h = pg.eval_on_selector('#view', 'e=>e.innerHTML')
            if 'Error al dibujar' in h:
                errores.append('Vista con error: %s' % v)
            if len(h) < 120:
                errores.append('Vista casi vacía: %s (%d bytes)' % (v, len(h)))
        for m in ['DEFINIR', 'MEDIR', 'ANALIZAR', 'MEJORAR', 'CONTROLAR']:
            pg.evaluate('(m)=>go("modulo",m)', m)
            pg.wait_for_timeout(120)

        # ── 4. escribir datos de prueba ────────────────────────────────
        pg.evaluate('''()=>{
          const p = ST.proj();
          p.nombre='Reducción de accidentalidad en planta';
          p.empresa='Manufacturas del Caribe S.A.S.';
          p.arl='ARL Positiva'; p.area='Ensamble'; p.lider='Ing. Ariel Ramírez';
          // llenar toda tabla con datos plausibles y todo campo de texto
          TOOLS.forEach(t=>{
            const d = ST.d(t.id);
            (t.blocks||[]).forEach(b=>{
              if(b.type==='fields'){
                (typeof b.items==='function'? b.items(d,ST.ctx()) : b.items||[]).forEach(f=>{
                  if(f.calc||f.ro) return;
                  if(f.input==='number') d[f.k]= Math.round(Math.random()*90)+10;
                  else if(f.input==='date') d[f.k]= '2026-03-15';
                  else if(f.input==='select'){ const o=(typeof f.opts==='function'?f.opts():f.opts)||[];
                    if(o.length) d[f.k]= (o[0]&&o[0].v!==undefined)?o[0].v:o[0]; }
                  else if(f.input==='check') d[f.k]=true;
                  else d[f.k]='Dato de prueba '+f.k;
                });
              }
              if(b.type==='grid'){
                const cols = typeof b.cols==='function'? b.cols(d,ST.ctx()) : (b.cols||[]);
                const rows=[];
                for(let i=0;i<Math.max(3,b.min||0);i++){
                  const r={_id:'t'+i};
                  cols.forEach(c=>{
                    if(c.calc||c.ro) return;
                    if(c.input==='number') r[c.k]= Math.round(Math.random()*40)+5;
                    else if(c.input==='date') r[c.k]= i%2? '2026-04-0'+(i+1) : '2026-05-1'+i;
                    else if(c.input==='select'){ const o=(typeof c.opts==='function'?c.opts():c.opts)||[];
                      if(o.length) r[c.k]= (o[0]&&o[0].v!==undefined)?o[0].v:o[0]; }
                    else r[c.k]= c.label.slice(0,14)+' '+(i+1);
                  });
                  rows.push(r);
                }
                d[b.key]=rows;
              }
              if(b.type==='a3'){
                (typeof b.areas==='function'? b.areas(d,ST.ctx()) : b.areas||[]).forEach(a=>{
                  if(a.kind==='text'||!a.kind) d[a.k]='Contenido A3 de prueba para '+a.label;
                });
              }
              if(b.type==='matrix'){
                d[b.key]=d[b.key]||{};
                const rws=(typeof b.rows==='function'?b.rows(d,ST.ctx()):b.rows)||[];
                const cls=(typeof b.cols==='function'?b.cols(d,ST.ctx()):b.cols)||[];
                rws.forEach(rw=>{ const rk=rw.k!==undefined?rw.k:rw;
                  cls.forEach(c=>{ const key=rk+'.'+c.k;
                    if(!(b.calc&&b.calc[key])) d[b.key][key]= c.input==='number'? 25 : 'X'; }); });
              }
            });
          });
          ST.save(true);
        }''')
        pg.wait_for_timeout(400)

        # redibujar todo con datos (detecta NaN/Infinity y errores de calc)
        rotas2 = []
        for tid in ids:
            pg.evaluate('(id)=>{const t=TOOL_BY_ID[id]; go("tool",t.mod,t.id);}', tid)
            pg.wait_for_timeout(35)
            h = pg.eval_on_selector('#view', 'e=>e.innerHTML')
            if 'Error al dibujar' in h:
                rotas2.append(tid)
            for bad in ['NaN', 'Infinity', 'undefined%', '[object Object]']:
                if bad in h:
                    avisos.append('%s muestra "%s"' % (tid, bad))
        if rotas2:
            errores.append('Con datos, fallan: %s' % rotas2)
        print('Segundo recorrido (con datos) · con error: %d' % len(rotas2))

        pg.evaluate('()=>go("bi")')
        pg.wait_for_timeout(600)
        bih = pg.eval_on_selector('#view', 'e=>e.innerHTML')
        print('Tablero BI: %d bytes' % len(bih))
        if len(bih) < 2000:
            errores.append('El tablero BI quedó demasiado pequeño (%d bytes)' % len(bih))

        # ── 5. exportaciones ───────────────────────────────────────────
        def bajar(js, nombre):
            try:
                with pg.expect_download(timeout=90000) as dl:
                    pg.evaluate(js)
                d = dl.value
                path = os.path.join(DL, d.suggested_filename)
                d.save_as(path)
                print('   ⤓ %-46s %8.0f KB' % (d.suggested_filename, os.path.getsize(path) / 1024))
                return path
            except Exception as ex:
                errores.append('Falló la descarga %s: %s' % (nombre, ex))
                return None

        print('Exportaciones:')
        xl1 = bajar('()=>XL.exportModule("DEFINIR")', 'DEFINIR.xlsx')
        xl2 = bajar('()=>XL.exportModule("MEDIR")', 'MEDIR.xlsx')
        xl3 = bajar('()=>XL.exportModule("CONTROLAR")', 'CONTROLAR.xlsx')
        zipall = bajar('()=>XL.exportAllTemplates()', 'todos.zip')
        xlsm = bajar('()=>GEN.exportConsolidado(true)', 'TODO EN UNO.xlsm')
        xlsx = bajar('()=>GEN.exportConsolidado(false)', 'TODO EN UNO.xlsx')
        csvz = bajar('()=>GEN.exportCSV()', 'csv.zip')
        bajar('()=>exportJSON()', 'respaldo.json')

        b.close()

    # ── 6. validar los archivos generados ──────────────────────────────
    print('\nValidación de archivos:')
    import openpyxl, warnings
    warnings.filterwarnings('ignore')
    for p in sorted(glob.glob(os.path.join(DL, '*.xls*'))):
        try:
            wb = openpyxl.load_workbook(p)
            llenas = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if c.value not in (None, ''):
                            llenas += 1
            print('   ✓ %-46s %2d hojas · %5d celdas con valor' %
                  (os.path.basename(p), len(wb.sheetnames), llenas))
            if llenas < 50:
                errores.append('%s casi vacío (%d celdas)' % (os.path.basename(p), llenas))
        except Exception as ex:
            errores.append('openpyxl no pudo abrir %s: %s' % (os.path.basename(p), ex))
            print('   ✗ %s → %s' % (os.path.basename(p), ex))

    for p in glob.glob(os.path.join(DL, '*.xlsm')):
        try:
            from oletools.olevba import VBA_Parser
            vp = VBA_Parser(p)
            mods = [v for (f, s, v, c) in vp.extract_macros()]
            print('   ✓ macros en %s → %s' % (os.path.basename(p), mods))
            if not mods:
                errores.append('El .xlsm no contiene macros')
        except Exception as ex:
            errores.append('Fallo al leer macros: %s' % ex)

    for p in glob.glob(os.path.join(DL, '*.zip')):
        try:
            z = zipfile.ZipFile(p)
            bad = z.testzip()
            print('   ✓ %-46s %d archivos%s' % (os.path.basename(p), len(z.namelist()),
                                                '' if bad is None else ' CORRUPTO:' + bad))
            if bad:
                errores.append('ZIP corrupto: %s' % p)
        except Exception as ex:
            errores.append('ZIP ilegible %s: %s' % (p, ex))

    # ── informe ────────────────────────────────────────────────────────
    print('\n' + '=' * 72)
    if avisos:
        vistos = {}
        for a in avisos:
            vistos[a[:110]] = vistos.get(a[:110], 0) + 1
        print('AVISOS (%d):' % len(avisos))
        for k, v in sorted(vistos.items(), key=lambda x: -x[1])[:25]:
            print('   [%dx] %s' % (v, k))
    if errores:
        print('\nERRORES (%d):' % len(errores))
        for e in errores[:40]:
            print('   ✗ %s' % str(e)[:400])
        return 1
    print('SIN ERRORES ✓')
    return 0


if __name__ == '__main__':
    try:
        sys.exit(run())
    except Exception:
        traceback.print_exc()
        sys.exit(2)
